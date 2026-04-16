from __future__ import annotations

import json
import sqlite3
from types import SimpleNamespace

import pytest


def _make_header():
    return SimpleNamespace(
        vessel_name="Geo Vessel",
        sections={
            "GEOMETRY": {
                "Offset(MBES)": "MBES,1.0,2.0,3.0",
            },
            "CALIB": {
                "StaticRoll": "roll,0.0",
                "StaticPitch": "pitch,0.0",
            },
        },
    )


def _create_offsetmanager_db(db_path):
    with sqlite3.connect(db_path) as conn:
        conn.execute(
            """
            CREATE TABLE sensor_offsets (
                id INTEGER PRIMARY KEY,
                config_id INTEGER,
                sensor_name TEXT,
                sensor_type TEXT,
                x_offset REAL,
                y_offset REAL,
                z_offset REAL,
                roll_offset REAL,
                pitch_offset REAL,
                heading_offset REAL,
                latency REAL
            )
            """
        )
        conn.execute(
            """
            INSERT INTO sensor_offsets
            (id, config_id, sensor_name, sensor_type, x_offset, y_offset, z_offset, roll_offset, pitch_offset, heading_offset, latency)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (1, 7, "MBES", "MBES Transducer", 1.0, 2.0, 3.0, 0.0, 0.0, 0.0, 0.0),
        )
        conn.commit()


def test_validate_offsets_attaches_provenance(monkeypatch, tmp_path):
    from mbes_qc import offset_validator

    db_path = tmp_path / "offsets.db"
    _create_offsetmanager_db(db_path)
    monkeypatch.setattr(offset_validator, "read_pds_header", lambda *a, **k: _make_header())
    monkeypatch.setattr(offset_validator, "read_pds_binary", lambda *a, **k: None)

    result = offset_validator.validate_offsets(
        "dummy.pds",
        offsetmanager_db=str(db_path),
        config_id=7,
        check_data=False,
    )

    assert result.overall == "PASS"
    assert result.provenance["om"]["decision"]["source"] == "db"
    assert result.provenance["om"]["decision"]["path"] == str(db_path)
    assert result.provenance["om"]["semantic"]["state"] == "verified"
    assert isinstance(result.provenance["om"]["semantic"]["checks"], list)
    assert result.provenance["om"]["resolution_chain"][0]["source"] == "db"


def test_result_json_reloads_with_provenance(monkeypatch, tmp_path):
    from desktop.services import analysis_service, data_service
    from desktop.services.insight_service import build_module_story, build_project_context
    from mbes_qc import offset_validator

    db_path = tmp_path / "offsets.db"
    _create_offsetmanager_db(db_path)
    monkeypatch.setattr(offset_validator, "read_pds_header", lambda *a, **k: _make_header())
    monkeypatch.setattr(offset_validator, "read_pds_binary", lambda *a, **k: None)

    validation = offset_validator.validate_offsets(
        "dummy.pds",
        offsetmanager_db=str(db_path),
        config_id=7,
        check_data=False,
    )
    serialized = analysis_service.serialize_full_qc_result(SimpleNamespace(offset_validation=validation))

    assert serialized["provenance"]["om"]["decision"]["source"] == "db"
    assert serialized["offset_validation"]["provenance"]["om"]["semantic"]["state"] == "verified"
    assert serialized["provenance_summary"]["source"] == "db"
    assert serialized["provenance_summary"]["semantic_state"] == "verified"
    assert serialized["provenance_manifest"]["type"] == "mbesqc.provenance-manifest"
    assert serialized["provenance_manifest"]["summary"] == serialized["provenance_summary"]

    monkeypatch.setattr(data_service, "_DB_PATH", tmp_path / "mbesqc_desktop.db", raising=False)
    data_service._local.conn = None
    data_service.DataService.init_db()
    project_id = data_service.DataService.create_project("Demo", vessel="Geo Vessel")
    result_id = data_service.DataService.create_qc_result(None, project_id)
    data_service.DataService.update_qc_result(
        result_id,
        status="done",
        result_json=json.dumps(serialized, ensure_ascii=False),
    )

    row = data_service.DataService.get_qc_result(result_id)
    assert row is not None
    assert row["provenance_summary"]["source"] == "db"
    assert row["provenance_summary"]["semantic_state"] == "verified"
    assert row["provenance_summary"] == serialized["provenance_summary"]
    assert row["provenance_manifest"]["type"] == "mbesqc.provenance-manifest"
    assert row["provenance_manifest"]["summary"] == row["provenance_summary"]
    assert row["result_payload"]["provenance_summary"]["semantic_checks_total"] >= 1
    assert row["result_payload"]["provenance_manifest"]["source"] == "db"
    assert row["result_payload"]["provenance"]["om"]["decision"]["source"] == "db"
    assert row["result_payload"]["offset_validation"]["provenance"]["om"]["semantic"]["state"] == "verified"

    project = data_service.DataService.get_project(project_id)
    context = build_project_context(
        project,
        latest_result=row,
        file_counts={"pds_count": 0, "gsf_count": 0, "hvf_count": 1},
    )
    assert "provenance db / verified" in context["snapshot_text"]
    assert "Provenance" in context["export_text"]

    story = build_module_story("offset", row["result_payload"])
    assert "db / verified" in story["current_reading"]


def test_serialize_full_qc_result_reuses_persisted_typed_provenance(monkeypatch, tmp_path):
    from desktop.services import analysis_service, data_service
    from mbes_qc import offset_validator

    db_path = tmp_path / "offsets.db"
    _create_offsetmanager_db(db_path)
    monkeypatch.setattr(offset_validator, "read_pds_header", lambda *a, **k: _make_header())
    monkeypatch.setattr(offset_validator, "read_pds_binary", lambda *a, **k: None)

    validation = offset_validator.validate_offsets(
        "dummy.pds",
        offsetmanager_db=str(db_path),
        config_id=7,
        check_data=False,
    )
    stored_summary = {
        "has_data": True,
        "source": "persisted",
        "mode": "sqlite",
        "path": str(db_path),
        "semantic_state": "verified",
        "semantic_hint": "typed summary already persisted",
        "semantic_checks_total": 1,
        "semantic_checks_passed": 1,
        "request_fallback_enabled": False,
        "project_fallback_enabled": False,
        "project_fallback_configured": False,
        "fallback_scope": "none",
    }
    stored_manifest = {
        "type": "mbesqc.provenance-manifest",
        "version": 1,
        "summary": stored_summary,
        "has_data": True,
        "source": "persisted",
        "mode": "sqlite",
        "path": str(db_path),
        "semantic_state": "verified",
        "semantic_hint": "typed summary already persisted",
        "semantic_checks_total": 1,
        "semantic_checks_passed": 1,
        "request_fallback_enabled": False,
        "project_fallback_enabled": False,
        "project_fallback_configured": False,
        "fallback_scope": "none",
    }

    monkeypatch.setattr(
        data_service.DataService,
        "extract_provenance_summary",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("unexpected raw provenance summary derivation")),
    )
    monkeypatch.setattr(
        data_service.DataService,
        "extract_provenance_manifest",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(AssertionError("unexpected raw provenance manifest derivation")),
    )

    serialized = analysis_service.serialize_full_qc_result(
        SimpleNamespace(
            offset_validation=validation,
            provenance_summary=stored_summary,
            provenance_manifest=stored_manifest,
        )
    )

    assert serialized["provenance_summary"] == stored_summary
    assert serialized["provenance_manifest"] == stored_manifest


def test_export_worker_preserves_provenance_sheet(monkeypatch, tmp_path):
    from desktop.services import analysis_service, data_service
    from desktop.services.export_service import ExportWorker
    from mbes_qc import offset_validator

    db_path = tmp_path / "offsets.db"
    _create_offsetmanager_db(db_path)
    monkeypatch.setattr(offset_validator, "read_pds_header", lambda *a, **k: _make_header())
    monkeypatch.setattr(offset_validator, "read_pds_binary", lambda *a, **k: None)

    validation = offset_validator.validate_offsets(
        "dummy.pds",
        offsetmanager_db=str(db_path),
        config_id=7,
        check_data=False,
    )
    serialized = analysis_service.serialize_full_qc_result(SimpleNamespace(offset_validation=validation))

    monkeypatch.setattr(data_service, "_DB_PATH", tmp_path / "mbesqc_desktop.db", raising=False)
    data_service._local.conn = None
    data_service.DataService.init_db()
    project_id = data_service.DataService.create_project("Demo", vessel="Geo Vessel")
    result_id = data_service.DataService.create_qc_result(None, project_id)
    data_service.DataService.update_qc_result(
        result_id,
        status="done",
        result_json=json.dumps(serialized, ensure_ascii=False),
    )

    output_path = tmp_path / "report.xlsx"
    worker = ExportWorker(project_id, "excel", str(output_path))
    _, _, _, _, latest_data, _, _, _, _, _, _, _, _, _, provenance = worker._load_export_context()

    assert provenance["summary"]["source"] == "db"
    assert provenance["summary"]["semantic_state"] == "verified"
    assert provenance["decision"]["source"] == "db"
    assert provenance["semantic"]["state"] == "verified"
    assert provenance["semantic_checks"]
    assert provenance["manifest"]["type"] == "mbesqc.provenance-manifest"
    assert latest_data["offset_validation"]["provenance"]["om"]["semantic"]["state"] == "verified"

    worker._export_excel()

    import openpyxl

    wb = openpyxl.load_workbook(output_path)
    assert "Provenance" in wb.sheetnames
    ws = wb["Provenance"]

    cell_values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None]
    assert any("Persisted summary" in value for value in cell_values)
    assert any("Decision source" in value for value in cell_values)
    assert any("db" in value for value in cell_values)
    assert any("Semantic checks" in value for value in cell_values)
    assert any("Raw provenance JSON" in value for value in cell_values)


def test_export_worker_redacts_absolute_provenance_paths(monkeypatch, tmp_path):
    from desktop.services import analysis_service, data_service
    from desktop.services.export_service import ExportWorker
    from mbes_qc import offset_validator

    db_path = tmp_path / "offsets.db"
    _create_offsetmanager_db(db_path)
    monkeypatch.setattr(offset_validator, "read_pds_header", lambda *a, **k: _make_header())
    monkeypatch.setattr(offset_validator, "read_pds_binary", lambda *a, **k: None)

    validation = offset_validator.validate_offsets(
        "dummy.pds",
        offsetmanager_db=str(db_path),
        config_id=7,
        check_data=False,
    )
    serialized = analysis_service.serialize_full_qc_result(SimpleNamespace(offset_validation=validation))
    sensitive_path = r"E:\Sensitive\Boundary\offsets.db"
    serialized["provenance_summary"]["path"] = sensitive_path
    serialized["provenance_manifest"]["path"] = sensitive_path
    serialized["provenance"]["om"]["decision"]["path"] = sensitive_path
    serialized["provenance"]["om"]["decision"]["filepath"] = sensitive_path
    serialized["provenance"]["om"]["decision"]["sourceFilePath"] = sensitive_path
    serialized["provenance"]["om"]["decision"]["db_path"] = sensitive_path
    serialized["provenance"]["om"]["fallback_candidate"]["path"] = sensitive_path
    serialized["provenance"]["om"]["fallback_candidate"]["source_path"] = sensitive_path
    serialized["provenance"]["om"]["fallback_candidate"]["source_filepath"] = sensitive_path
    serialized["provenance"]["om"]["fallback_candidate"]["sourceFilePath"] = sensitive_path
    serialized["provenance"]["om"]["semantic"]["path"] = sensitive_path
    serialized["provenance"]["om"]["resolution_chain"][0]["path"] = sensitive_path
    serialized["provenance"]["om"]["resolution_chain"][0]["file_path"] = sensitive_path
    serialized["provenance"]["om"]["resolution_chain"][0]["source_filepath"] = sensitive_path
    serialized["provenance"]["om"]["resolution_chain"][0]["filePath"] = sensitive_path
    serialized["provenance"]["om"]["resolution_chain"][1]["path"] = sensitive_path
    serialized["offset_validation"]["provenance"]["om"]["decision"]["path"] = sensitive_path

    monkeypatch.setattr(data_service, "_DB_PATH", tmp_path / "mbesqc_desktop.db", raising=False)
    data_service._local.conn = None
    data_service.DataService.init_db()
    project_id = data_service.DataService.create_project("Demo", vessel="Geo Vessel")
    result_id = data_service.DataService.create_qc_result(None, project_id)
    data_service.DataService.update_qc_result(
        result_id,
        status="done",
        result_json=json.dumps(serialized, ensure_ascii=False),
    )

    output_path = tmp_path / "redacted_report.xlsx"
    worker = ExportWorker(project_id, "excel", str(output_path))
    worker._export_excel()

    import openpyxl

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Provenance"]
    cell_values = [str(cell.value) for row in ws.iter_rows() for cell in row if cell.value is not None]

    assert not any("Sensitive" in value for value in cell_values)
    assert any("offsets.db" in value for value in cell_values)
    assert any("\"sourceFilePath\": \"offsets.db\"" in value for value in cell_values)
    assert any("\"source_filepath\": \"offsets.db\"" in value for value in cell_values)


def test_export_worker_cleans_temp_chart_files_when_save_fails(monkeypatch, tmp_path):
    from pathlib import Path

    from desktop.services import chart_renderer, export_service
    from desktop.services.export_service import ExportWorker

    worker = ExportWorker(7, "excel", str(tmp_path / "report.xlsx"))

    worker._load_export_context = lambda: (
        {"name": "Demo", "vessel": "Geo Vessel"},
        [],
        [],
        {"status": "done"},
        {"any": True},
        {
            "snapshot_text": "Snapshot",
            "flow": "Flow",
            "offset_text": "Offset",
            "readiness_text": "Ready",
        },
        {
            "headline": "Headline",
            "body": "Body",
            "next_steps": ["Next"],
        },
        [
            {
                "module": "Offset",
                "status": "PASS",
                "importance": "Why it matters",
                "current_reading": "Reading",
                "next_steps": ["Check"],
            }
        ],
        [
            {
                "snapshot": "Snapshot",
                "score": "1.0",
                "grade": "A",
                "status": "done",
                "completed": "2026-04-13",
            }
        ],
        [],
        [],
        {"headline": "History", "body": "History body"},
        {"body": "Diff body"},
        {
            "current_state": ["state"],
            "recommendations": ["rec"],
        },
        {
            "has_data": True,
            "summary": {
                "has_data": True,
                "source": "db",
                "semantic_state": "verified",
                "semantic_checks_total": 1,
                "semantic_checks_passed": 1,
            },
            "decision": {
                "source": "db",
                "path": "offsets.db",
            },
            "semantic": {
                "state": "verified",
                "hint": "ok",
                "checks": [{"passed": True}],
            },
            "resolution_chain": [{"source": "db"}],
        },
    )

    png_bytes = (
        b"\x89PNG\r\n\x1a\n"
        b"\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01"
        b"\x08\x02\x00\x00\x00\x90wS\xde"
        b"\x00\x00\x00\nIDATx\x9cc`\x00\x00\x00\x02\x00\x01"
        b"\xe2!\xbc3\x00\x00\x00\x00IEND\xaeB`\x82"
    )
    created_paths: list[Path] = []
    real_named_tempfile = export_service.tempfile.NamedTemporaryFile

    def _fake_named_tempfile(*args, **kwargs):
        kwargs = dict(kwargs)
        kwargs["dir"] = str(tmp_path)
        handle = real_named_tempfile(*args, **kwargs)
        created_paths.append(Path(handle.name))
        return handle

    def _fake_render_qc_radar(*_args, **_kwargs):
        return png_bytes

    def _fail_save(self, *_args, **_kwargs):
        raise RuntimeError("save failed")

    monkeypatch.setattr(chart_renderer, "render_qc_radar", _fake_render_qc_radar)
    monkeypatch.setattr(export_service.tempfile, "NamedTemporaryFile", _fake_named_tempfile)

    import openpyxl.workbook.workbook as workbook_module

    monkeypatch.setattr(workbook_module.Workbook, "save", _fail_save)

    with pytest.raises(RuntimeError, match="save failed"):
        worker._export_excel()

    assert created_paths, "expected a chart temp file to be created"
    assert all(not path.exists() for path in created_paths)
    assert not list(tmp_path.glob("mbesqc_radar_*.png"))


def test_download_payload_includes_provenance_manifest(monkeypatch, tmp_path):
    from desktop.services import analysis_service, data_service
    import web_app
    from mbes_qc import offset_validator

    db_path = tmp_path / "offsets.db"
    _create_offsetmanager_db(db_path)
    monkeypatch.setattr(offset_validator, "read_pds_header", lambda *a, **k: _make_header())
    monkeypatch.setattr(offset_validator, "read_pds_binary", lambda *a, **k: None)

    validation = offset_validator.validate_offsets(
        "dummy.pds",
        offsetmanager_db=str(db_path),
        config_id=7,
        check_data=False,
    )
    serialized = analysis_service.serialize_full_qc_result(SimpleNamespace(offset_validation=validation))
    serialized["provenance"]["om"]["decision"]["uri"] = "file:///E:/Sensitive/Boundary/offsets.db"
    serialized["provenance"]["om"]["decision"]["filepath"] = r"E:\Sensitive\Boundary\offsets.db"
    serialized["provenance"]["om"]["decision"]["sourceFilePath"] = r"E:\Sensitive\Boundary\offsets.db"
    serialized["provenance"]["om"]["decision"]["db_path"] = r"E:\Sensitive\Boundary\offsets.db"
    serialized["provenance"]["om"]["fallback_candidate"]["uri"] = "https://offsetmanager.example/api/configs"
    serialized["provenance"]["om"]["fallback_candidate"]["source_path"] = r"E:\Sensitive\Boundary\offsets.db"
    serialized["provenance"]["om"]["fallback_candidate"]["source_filepath"] = r"E:\Sensitive\Boundary\offsets.db"
    serialized["provenance"]["om"]["fallback_candidate"]["sourceFilePath"] = r"E:\Sensitive\Boundary\offsets.db"
    serialized["provenance"]["om"]["resolution_chain"][0]["uri"] = "file:///E:/Sensitive/Boundary/offsets.db"
    serialized["provenance"]["om"]["resolution_chain"][0]["file_path"] = r"E:\Sensitive\Boundary\offsets.db"
    serialized["provenance"]["om"]["resolution_chain"][0]["source_filepath"] = r"E:\Sensitive\Boundary\offsets.db"
    serialized["provenance"]["om"]["resolution_chain"][0]["filePath"] = r"E:\Sensitive\Boundary\offsets.db"
    serialized["provenance"]["om"]["resolution_chain"][1]["uri"] = "https://offsetmanager.example/api/configs"

    monkeypatch.setattr(data_service, "_DB_PATH", tmp_path / "mbesqc_desktop.db", raising=False)
    data_service._local.conn = None
    data_service.DataService.init_db()

    monkeypatch.setattr(
        web_app,
        "_jobs",
        {
            1: {
                "status": "done",
                "result": serialized,
            }
        },
        raising=False,
    )

    payload = web_app._build_job_download_payload(web_app._jobs[1])
    payload_text = json.dumps(payload, ensure_ascii=False)
    assert payload["provenance_summary"]["source"] == "db"
    assert payload["provenance_manifest"]["type"] == "mbesqc.provenance-manifest"
    assert payload["provenance_manifest"]["summary"] == payload["provenance_summary"]
    assert payload["provenance"]["om"]["decision"]["uri"] == "<redacted:uri>"
    assert payload["provenance"]["om"]["decision"]["filepath"] == "offsets.db"
    assert payload["provenance"]["om"]["decision"]["sourceFilePath"] == "offsets.db"
    assert payload["provenance"]["om"]["decision"]["db_path"] == "offsets.db"
    assert payload["provenance"]["om"]["fallback_candidate"]["uri"] == "<redacted:uri>"
    assert payload["provenance"]["om"]["fallback_candidate"]["source_path"] == "offsets.db"
    assert payload["provenance"]["om"]["fallback_candidate"]["source_filepath"] == "offsets.db"
    assert payload["provenance"]["om"]["fallback_candidate"]["sourceFilePath"] == "offsets.db"
    assert payload["provenance"]["om"]["resolution_chain"][0]["source_filepath"] == "offsets.db"
    assert payload["provenance"]["om"]["resolution_chain"][0]["filePath"] == "offsets.db"
    assert "Sensitive" not in payload_text
